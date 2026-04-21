"""通用 Excel 读取器，支持 .xlsx 和 .xls 格式。"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass
class SheetData:
    """单个 Sheet 的原始数据。"""

    sheet_name: str
    rows: list[list[Any]]
    merged_ranges: list[str]
    max_row: int
    max_col: int

    def __post_init__(self):
        self._merge_top_left: dict[tuple[int, int], tuple[int, int]] = {}
        for rng in self.merged_ranges:
            parts = rng.split(":")
            if len(parts) != 2:
                continue
            r1, c1 = _cell_ref_to_index(parts[0])
            r2, c2 = _cell_ref_to_index(parts[1])
            top_left_value = (
                self.rows[r1][c1]
                if r1 < len(self.rows) and c1 < len(self.rows[r1])
                else None
            )
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    self._merge_top_left[(r, c)] = (r1, c1)
                    if (r, c) != (r1, c1):
                        if r < len(self.rows) and c < len(self.rows[r]):
                            self.rows[r][c] = top_left_value

    def get_merged_value(self, row: int, col: int) -> Any:
        if (row, col) in self._merge_top_left:
            tr, tc = self._merge_top_left[(row, col)]
            if tr < len(self.rows) and tc < len(self.rows[tr]):
                return self.rows[tr][tc]
        if row < len(self.rows) and col < len(self.rows[row]):
            return self.rows[row][col]
        return None


def _col_letter_to_index(letter: str) -> int:
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def _cell_ref_to_index(ref: str) -> tuple[int, int]:
    col_str = ""
    row_str = ""
    for ch in ref:
        if ch.isdigit():
            row_str += ch
        else:
            col_str += ch
    return _col_letter_to_index(col_str), int(row_str) - 1


class ExcelReader:
    """统一 Excel 读取器，支持 .xlsx 和 .xls 格式。"""

    def __init__(self, file_path: str, max_rows_per_sheet: int = 200):
        self._path = Path(file_path)
        self._max_rows = max_rows_per_sheet
        if not self._path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        self._ext = self._path.suffix.lower()
        if self._ext not in (".xlsx", ".xls"):
            raise ValueError(f"Unsupported format: {self._ext}. Only .xlsx and .xls are supported.")

    @property
    def sheet_names(self) -> list[str]:
        if self._ext == ".xlsx":
            return self._sheet_names_xlsx()
        return self._sheet_names_xls()

    def read_sheet(self, sheet_name: str) -> SheetData:
        if self._ext == ".xlsx":
            return self._read_sheet_xlsx(sheet_name)
        return self._read_sheet_xls(sheet_name)

    def read_all_sheets(self) -> dict[str, SheetData]:
        return {name: self.read_sheet(name) for name in self.sheet_names}

    def _sheet_names_xlsx(self) -> list[str]:
        import openpyxl
        wb = openpyxl.load_workbook(str(self._path), read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names

    def _sheet_names_xls(self) -> list[str]:
        import xlrd
        wb = xlrd.open_workbook(str(self._path))
        return [s.name for s in wb.sheets()]

    def _read_sheet_xlsx(self, sheet_name: str) -> SheetData:
        import openpyxl
        wb = openpyxl.load_workbook(str(self._path), data_only=True)
        ws = wb[sheet_name]
        merged = [str(m) for m in ws.merged_cells.ranges]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= self._max_rows:
                break
            rows.append(list(row))
        wb.close()
        return SheetData(sheet_name=sheet_name, rows=rows, merged_ranges=merged,
                         max_row=len(rows), max_col=max((len(r) for r in rows), default=0))

    def _read_sheet_xls(self, sheet_name: str) -> SheetData:
        import xlrd
        from openpyxl.utils import get_column_letter
        wb = xlrd.open_workbook(str(self._path), formatting_info=True)
        ws = wb.sheet_by_name(sheet_name)
        nrows = min(ws.nrows, self._max_rows)
        rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(nrows)]
        merged = []
        for r1, r2, c1, c2 in ws.merged_cells:
            if r1 < self._max_rows:
                merged.append(f"{get_column_letter(c1+1)}{r1+1}:{get_column_letter(c2)}{r2}")
        return SheetData(sheet_name=sheet_name, rows=rows, merged_ranges=merged,
                         max_row=nrows, max_col=ws.ncols)
